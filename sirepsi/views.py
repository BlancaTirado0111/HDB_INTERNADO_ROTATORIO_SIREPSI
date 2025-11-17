# sirepsi/views.py
from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re
import csv
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlencode

from django.core.cache import cache
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

# ================== DEPENDENCIAS OPCIONALES ==================

# PDF (ReportLab)
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    HAS_REPORTLAB = True
except Exception:
    HAS_REPORTLAB = False

# XLSX (openpyxl)
try:
    import openpyxl
    from openpyxl.styles import Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

from .dbutils import (
    query_bdfarmacia,
    kardex_encabezado,
    kardex_detalle,
    prescripciones_detalle,
)

# ============================================================
#                   HOME / UTILES
# ============================================================

DEFAULT_ALMACEN: int = 31  # fallback

def listar_almacenes() -> List[Dict[str, Any]]:
    """
    Devuelve [{'id': 31, 'nombre': 'ALMACEN ...'}, ...]
    Prueba fa_almacen e in_almacen con variantes comunes de nombres de columna.
    """
    probes = [
        ("dbo.fa_almacen", ("ALM_CODIGO", "ALM_DESCRIPCION")),
        ("dbo.fa_almacen", ("Alm_Codigo", "Alm_Nombre")),
        ("dbo.in_almacen", ("ALM_CODIGO", "ALM_DESCRIPCION")),
        ("dbo.in_almacen", ("Alm_Codigo", "Alm_Nombre")),
    ]
    for tabla, (col_id, col_nom) in probes:
        try:
            rows = query_bdfarmacia(
                f"""
                SELECT CAST({col_id} AS INT)  AS id,
                       RTRIM(ISNULL({col_nom},'')) AS nombre
                FROM {tabla}
                ORDER BY {col_id}
                """,
                None,
            )
            if rows:
                return rows
        except Exception:
            continue
    return []


def _resolver_almacen(get_dict, almacenes: List[Dict[str, Any]]) -> Tuple[int, str, bool]:
    """
    Resuelve el almacén desde GET['alm'].
    Devuelve: (alm_int, mensaje_corto, lista_ok)
    """
    alm_s = (get_dict.get("alm") or "").strip()
    try:
        alm_int = int(alm_s) if alm_s else None
    except Exception:
        alm_int = None

    lista_ok = bool(almacenes)
    if alm_int is None:
        # Si hay lista, intenta usar DEFAULT si existe; si no, el primero
        if lista_ok:
            ids = {a["id"] for a in almacenes}
            alm_int = DEFAULT_ALMACEN if DEFAULT_ALMACEN in ids else almacenes[0]["id"]
        else:
            alm_int = DEFAULT_ALMACEN

    msg = f"Usando almacén #{alm_int}"
    return alm_int, msg, lista_ok


def home(request):
    return render(request, "home.html")


def ping(request):
    return render(request, "base.html", {"content": "pong"})


def _page_window(page_obj, window: int = 2) -> List[Optional[int]]:
    current = page_obj.number
    total = page_obj.paginator.num_pages
    pages = {1, total}
    for p in range(current - window, current + window + 1):
        if 1 <= p <= total:
            pages.add(p)
    for p in range(2, min(2 + window, total)):
        pages.add(p)
    for p in range(max(total - window + 1, 1), total):
        pages.add(p)
    pages = sorted(pages)
    compact: List[Optional[int]] = []
    prev: Optional[int] = None
    for p in pages:
        if prev is not None and p != prev + 1:
            compact.append(None)
        compact.append(p)
        prev = p
    return compact


# ============================================================
#                   MOVIMIENTOS: LISTA DE MEDS
# ============================================================

PSYCH_LIST: List[Tuple[str, str]] = [
    ("N0501", "Alprazolam"),
    ("N0306", "Clonazepam"),
    ("N0312", "Clonazepam"),
    ("N0504", "Diazepam"),
    ("N0505", "Diazepam"),
    ("N0309", "Fenobarbital"),
    ("N0310", "Fenobarbital"),
    ("N0311", "Fenobarbital"),
    ("N0105", "Fentanilo con conservante"),
    ("N0106", "Fentanilo sin conservante"),
    ("N0511", "Midazolam"),
    ("N0206", "Morfina"),
    ("N0207", "Morfina (con o sin conservante)"),
    ("N0116", "Remifentanilo"),
]

MED_TYPES_CHOICES = [
    ("benzo", "Benzodiacepinas"),
    ("barbiturico", "Barbitúricos"),
    ("opioide", "Opioides"),
]

MED_TYPES_CODES = {
    "benzo": {"N0501", "N0306", "N0312", "N0504", "N0505", "N0511"},
    "barbiturico": {"N0309", "N0310", "N0311"},
    "opioide": {"N0105", "N0106", "N0206", "N0207", "N0116"},
}


def _psych_codes_sql_list() -> str:
    codes = [c for c, _ in PSYCH_LIST]
    return ", ".join(f"'{c}'" for c in codes)


def medicamentos(request):
    """
    Lista paginada (40 por página) con filtros por texto, laboratorio y presentación.
    """
    q = (request.GET.get("q") or "").strip()
    lab = (request.GET.get("lab") or "").strip()
    pres = (request.GET.get("pres") or "").strip()
    order = (request.GET.get("order") or "nombre").lower()

    base_where = f"m.MED_CODIFICACION IN ({_psych_codes_sql_list()})"

    sql_choices = f"""
        SELECT DISTINCT
            RTRIM(ISNULL(p.PRO_NOMBRE,''))  AS laboratorio,
            RTRIM(ISNULL(m.med_unidad,''))  AS presentacion
        FROM dbo.fa_medicamento m
        LEFT JOIN dbo.fa_proveedor p
               ON p.Emp_Codigo = m.emp_codigo
        WHERE {base_where}
    """
    choice_rows = query_bdfarmacia(sql_choices, None)

    laboratorios = sorted({
        (row.get("laboratorio") or "").strip()
        for row in choice_rows if (row.get("laboratorio") or "").strip()
    })
    presentaciones = sorted({
        (row.get("presentacion") or "").strip()
        for row in choice_rows if (row.get("presentacion") or "").strip()
    })

    sql = f"""
        SELECT
            m.MED_CODIGO         AS codigo,
            m.MED_CODIFICACION   AS codificacion,
            m.MED_COMERCIAL      AS nombre,
            m.med_generico       AS dci,
            m.med_concentracion  AS concentracion,
            m.med_unidad         AS presentacion,
            p.PRO_NOMBRE         AS laboratorio
        FROM dbo.fa_medicamento m
        LEFT JOIN dbo.fa_proveedor p
               ON p.Emp_Codigo = m.emp_codigo
        WHERE {base_where}
    """

    where_clauses: List[str] = []
    params: List[Any] = []

    if q:
        like = f"%{q}%"
        where_clauses.append(
            "("
            "m.MED_COMERCIAL    LIKE %s OR "
            "m.med_generico     LIKE %s OR "
            "m.MED_CODIGO       LIKE %s OR "
            "m.MED_CODIFICACION LIKE %s"
            ")"
        )
        params.extend([like, like, like, like])
    if lab:
        where_clauses.append("p.PRO_NOMBRE = %s")
        params.append(lab)
    if pres:
        where_clauses.append("m.med_unidad = %s")
        params.append(pres)
    if where_clauses:
        sql += " AND " + " AND ".join(where_clauses)

    order_map: Dict[str, List[str]] = {
        "nombre": ["m.MED_COMERCIAL", "m.med_generico"],
        "codigo": ["m.MED_CODIGO"],
        "codificacion": ["m.MED_CODIFICACION", "m.MED_COMERCIAL"],
        "lab": ["p.PRO_NOMBRE", "m.MED_COMERCIAL"],
    }
    order_parts = order_map.get(order, order_map["nombre"]) + ["m.MED_CODIGO"]
    seen, unique_order_parts = set(), []
    for part in order_parts:
        if part not in seen:
            unique_order_parts.append(part)
            seen.add(part)
    sql += " ORDER BY " + ", ".join(unique_order_parts)

    filas = query_bdfarmacia(sql, params or None)
    paginator = Paginator(filas, 40)
    page_obj = paginator.get_page(request.GET.get("page"))

    no_results_message = ""
    if (q or lab or pres) and paginator.count == 0:
        no_results_message = "No hay existencias para esta combinación de filtros."

    qs_params: Dict[str, str] = {}
    if q:
        qs_params["q"] = q
    if lab:
        qs_params["lab"] = lab
    if pres:
        qs_params["pres"] = pres
    if order:
        qs_params["order"] = order
    querystring = "&" + urlencode(qs_params) if qs_params else ""

    context = {
        "page_obj": page_obj,
        "pages_compact": _page_window(page_obj, window=2),
        "q": q,
        "lab": lab,
        "pres": pres,
        "order": order,
        "laboratorios": laboratorios,
        "presentaciones": presentaciones,
        "total": paginator.count,
        "no_results_message": no_results_message,
        "querystring": querystring,
    }
    return render(request, "medicamentos/lista.html", context)


# ============================================================
#                   MOVIMIENTOS (KARDEX)
# ============================================================

def meds_suggest(request):
    q = (request.GET.get("q") or "").strip().lower()
    results: List[Dict[str, str]] = []
    for code, name in PSYCH_LIST:
        if not q or (q in code.lower()) or (q in name.lower()):
            results.append({"code": code, "label": f"{code} — {name}"})
        if len(results) >= 20:
            break
    return JsonResponse({"results": results})


def _resolve_med_by_name_or_code(name_or_code: str) -> Optional[Dict[str, Any]]:
    q = (name_or_code or "").strip()
    if not q:
        return None

    pick_code: Optional[str] = None
    for code, name in PSYCH_LIST:
        if q.lower() in name.lower() or q.upper() == code.upper():
            pick_code = code
            break

    if pick_code:
        rows = query_bdfarmacia(
            """
            SELECT TOP(1)
              m.MED_CODIGO         AS med_codigo,
              m.MED_CODIFICACION   AS codificacion,
              m.MED_COMERCIAL      AS nombre,
              m.med_generico       AS dci,
              m.med_concentracion  AS concentracion,
              m.med_unidad         AS presentacion,
              p.PRO_NOMBRE         AS laboratorio
            FROM dbo.fa_medicamento m
            LEFT JOIN dbo.fa_proveedor p ON p.Emp_Codigo = m.emp_codigo
            WHERE m.MED_CODIFICACION = %s
            """,
            [pick_code],
        )
        if rows:
            return rows[0]

    rows = query_bdfarmacia(
        """
        SELECT TOP(1)
          m.MED_CODIGO         AS med_codigo,
          m.MED_CODIFICACION   AS codificacion,
          m.MED_COMERCIAL      AS nombre,
          m.med_generico       AS dci,
          m.med_concentracion  AS concentracion,
          m.med_unidad         AS presentacion,
          p.PRO_NOMBRE         AS laboratorio
        FROM dbo.fa_medicamento m
        LEFT JOIN dbo.fa_proveedor p ON p.Emp_Codigo = m.emp_codigo
        WHERE     m.MED_COMERCIAL LIKE %s
            OR    m.MED_GENERICO  LIKE %s
        ORDER BY LEN(m.MED_COMERCIAL), m.MED_COMERCIAL
        """,
        [f"%{q}%", f"%{q}%"],
    )
    return rows[0] if rows else None


def movimientos_kardex(request):
    """
    Vista Kardex con filtro de ALMACÉN.
    """
    q = (request.GET.get("q") or "").strip()
    desde = (request.GET.get("desde") or "").strip()
    hasta = (request.GET.get("hasta") or "").strip()

    almacenes = listar_almacenes()
    alm_int, alm_msg, alm_lista_ok = _resolver_almacen(request.GET, almacenes)

    ctx = {
        "q": q,
        "desde": desde,
        "hasta": hasta,
        "alm": alm_int,
        "almacenes": almacenes,
        "alm_msg": alm_msg,          # Mensaje corto
        "alm_lista_ok": alm_lista_ok,  # True si el <select> puede poblarse
        "encabezado": None,
        "rows": [],
        "error": "",
        "saldo_inicial": 0,
        "sum_ingresos": 0,
        "sum_egresos": 0,
        "saldo_final": 0,
        "total_movs": 0,
    }

    if not (q or desde or hasta):
        return render(request, "movimientos/kardex.html", ctx)

    if len(q) < 3:
        ctx["error"] = "Escribe al menos 3 letras del nombre del medicamento."
        return render(request, "movimientos/kardex.html", ctx)

    try:
        d_desde = datetime.strptime(desde, "%Y-%m-%d").date()
        d_hasta = datetime.strptime(hasta, "%Y-%m-%d").date()
    except Exception:
        ctx["error"] = "Fechas inválidas. Usa el selector de fechas."
        return render(request, "movimientos/kardex.html", ctx)

    if d_hasta < d_desde:
        ctx["error"] = "La fecha hasta debe ser mayor o igual a la fecha desde."
        return render(request, "movimientos/kardex.html", ctx)

    med = _resolve_med_by_name_or_code(q)
    if not med:
        ctx["error"] = "No se encontró un medicamento con ese nombre."
        return render(request, "movimientos/kardex.html", ctx)

    cache_key = f"kardex:{alm_int}:{med['med_codigo']}:{d_desde.isoformat()}:{d_hasta.isoformat()}"
    data = cache.get(cache_key)
    if data is None:
        enc = kardex_encabezado(med["med_codigo"])
        rows = kardex_detalle(
            med_codigo=med["med_codigo"],
            almacen=alm_int,
            fecha_desde=d_desde,
            fecha_hasta=d_hasta,
        )
        cache.set(cache_key, (enc, rows), 300)
    else:
        enc, rows = data

    sum_ing = sum(float(r.get("cantidad_ingreso") or 0) for r in rows)
    sum_egr = sum(float(r.get("cantidad_egreso") or 0) for r in rows)
    saldo_ini = float(rows[0]["saldo_anterior"]) if rows else 0.0
    saldo_fin = float(rows[-1]["saldo_actual"]) if rows else saldo_ini

    ctx.update(
        {
            "q": (enc["nombre_del_producto"] if enc else med["nombre"]),
            "encabezado": {
                "nombre": enc.get("nombre_del_producto", med["nombre"]) if enc else med["nombre"],
                "dci": enc.get("dci", med["dci"]) if enc else med["dci"],
                "concentracion": enc.get("concentracion", med["concentracion"]) if enc else med["concentracion"],
                "presentacion": enc.get("presentacion", med["presentacion"]) if enc else med["presentacion"],
                "laboratorio": enc.get("laboratorio") if enc else None,
            },
            "rows": rows,
            "saldo_inicial": saldo_ini,
            "sum_ingresos": sum_ing,
            "sum_egresos": sum_egr,
            "saldo_final": saldo_fin,
            "total_movs": len(rows),
        }
    )
    return render(request, "movimientos/kardex.html", ctx)


# ============================================================
#                   PRESCRIPCIONES
# ============================================================

def prescripciones(request):
    """
    Lista de prescripciones por medicamento y rango de fechas, con filtro de ALMACÉN.
    """
    q = (request.GET.get("q") or "").strip()
    medico = (request.GET.get("medico") or "").strip()
    desde = (request.GET.get("desde") or "").strip()
    hasta = (request.GET.get("hasta") or "").strip()

    almacenes = listar_almacenes()
    alm_int, alm_msg, alm_lista_ok = _resolver_almacen(request.GET, almacenes)

    ctx = {
        "q": q,
        "medico": medico,
        "desde": desde,
        "hasta": hasta,
        "alm": alm_int,
        "almacenes": almacenes,
        "alm_msg": alm_msg,
        "alm_lista_ok": alm_lista_ok,
        "encabezado": None,
        "rows": [],
        "error": "",
        "total_presc": 0,
        "pacientes_unicos": 0,
        "medicos_unicos": 0,
        "medicos_disponibles": [],
    }

    if not (q or desde or hasta):
        return render(request, "prescripciones/lista.html", ctx)

    if len(q) < 3:
        ctx["error"] = "Escribe al menos 3 letras del nombre del medicamento."
        return render(request, "prescripciones/lista.html", ctx)

    try:
        d_desde = datetime.strptime(desde, "%Y-%m-%d").date()
        d_hasta = datetime.strptime(hasta, "%Y-%m-%d").date()
    except Exception:
        ctx["error"] = "Fechas inválidas. Usa el selector de fechas."
        return render(request, "prescripciones/lista.html", ctx)

    if (d_hasta - d_desde).days > 31:
        ctx["error"] = "Por ahora el reporte de prescripciones permite rangos de hasta 31 días."
        return render(request, "prescripciones/lista.html", ctx)

    if d_hasta < d_desde:
        ctx["error"] = "La fecha hasta debe ser mayor o igual a la fecha desde."
        return render(request, "prescripciones/lista.html", ctx)

    med = _resolve_med_by_name_or_code(q)
    if not med:
        ctx["error"] = "No se encontró un medicamento con ese nombre."
        return render(request, "prescripciones/lista.html", ctx)

    cache_key = f"presc:{alm_int}:{med['med_codigo']}:{d_desde.isoformat()}:{d_hasta.isoformat()}"
    data = cache.get(cache_key)
    if data is None:
        enc = kardex_encabezado(med["med_codigo"])
        rows = prescripciones_detalle(
            med_codigo=med["med_codigo"],
            almacen=alm_int,
            fecha_desde=d_desde,
            fecha_hasta=d_hasta,
        )
        cache.set(cache_key, (enc, rows), 300)
    else:
        enc, rows = data

    medicos_disponibles = sorted(
        {
            (r.get("nombre_medico") or "").strip()
            for r in rows
            if (r.get("nombre_medico") or "").strip()
        }
    )

    rows_filtradas = rows
    if medico:
        m_ref = medico.strip().lower()
        rows_filtradas = [
            r
            for r in rows
            if (r.get("nombre_medico") or "").strip().lower() == m_ref
        ]

    total = len(rows_filtradas)
    pacientes = len(
        {
            (r.get("nombre_paciente") or "").strip()
            for r in rows_filtradas
            if (r.get("nombre_paciente") or "").strip()
        }
    )
    medicos = len(
        {
            (r.get("nombre_medico") or "").strip()
            for r in rows_filtradas
            if (r.get("nombre_medico") or "").strip()
        }
    )

    ctx.update(
        {
            "q": enc["nombre_del_producto"] if enc else med["nombre"],
            "encabezado": {
                "nombre": enc.get("nombre_del_producto", med["nombre"]) if enc else med["nombre"],
                "dci": enc.get("dci", med["dci"]) if enc else med["dci"],
                "concentracion": enc.get("concentracion", med["concentracion"]) if enc else med["concentracion"],
                "presentacion": enc.get("presentacion", med["presentacion"]) if enc else med["presentacion"],
                "laboratorio": enc.get("laboratorio") if enc else None,
            },
            "rows": rows_filtradas,
            "total_presc": total,
            "pacientes_unicos": pacientes,
            "medicos_unicos": medicos,
            "medicos_disponibles": medicos_disponibles,
        }
    )
    return render(request, "prescripciones/lista.html", ctx)


# ============================================================
#                   EXPORTS (CSV / XLSX / PDF)
# ============================================================

def _safe_filename(base: str) -> str:
    base = (base or "reporte").strip()
    base = re.sub(r"[^\w\-.]+", "_", base, flags=re.UNICODE)
    return base or "reporte"


def _resolver_kardex_dataset(request):
    """
    Arma el dataset para exportar, respetando el almacén.
    GET: q, desde, hasta, alm

    Devuelve:
      enc, rows, d_desde, d_hasta, alm_int, med
    """
    q = (request.GET.get("q") or "").strip()
    desde = (request.GET.get("desde") or "").strip()
    hasta = (request.GET.get("hasta") or "").strip()
    alm_raw = (request.GET.get("alm") or "").strip()

    if not q:
        raise ValueError("Falta el nombre del medicamento.")
    try:
        d_desde = datetime.strptime(desde, "%Y-%m-%d").date()
        d_hasta = datetime.strptime(hasta, "%Y-%m-%d").date()
    except Exception:
        raise ValueError("Fechas inválidas. Usa el formato YYYY-MM-DD.")

    med = _resolve_med_by_name_or_code(q)
    if not med:
        raise ValueError("No se encontró el medicamento.")

    try:
        alm_int = int(alm_raw) if alm_raw else DEFAULT_ALMACEN
    except Exception:
        alm_int = DEFAULT_ALMACEN

    enc = kardex_encabezado(med["med_codigo"]) or {}
    rows = kardex_detalle(
        med_codigo=med["med_codigo"],
        almacen=alm_int,
        fecha_desde=d_desde,
        fecha_hasta=d_hasta,
    )
    return enc, rows, d_desde, d_hasta, alm_int, med


def _build_med_label(enc: Dict[str, Any], med: Dict[str, Any]) -> str:
    """
    Construye 'ALPRAZOLAM_N0501' usando nombre + codificación.
    """
    med_name = enc.get("nombre_del_producto") or med.get("nombre") or "medicamento"
    med_codif = enc.get("codificacion") or med.get("codificacion") or ""
    if med_codif:
        label = f"{med_name}_{med_codif}"
    else:
        label = med_name
    return _safe_filename(label)


def export_kardex_csv(request):
    try:
        enc, rows, d_desde, d_hasta, alm, med = _resolver_kardex_dataset(request)
    except Exception as e:
        return HttpResponse(
            f"Error: {e}", status=400, content_type="text/plain; charset=utf-8"
        )

    base = _build_med_label(enc, med)
    filename = f"KARDEX_{base}_DEL_{d_desde}_AL_{d_hasta}.csv"

    resp = HttpResponse(content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp.write("\ufeff")  # BOM
    w = csv.writer(resp, lineterminator="\r\n")

    w.writerow(["KARDEX"])
    w.writerow([f"Producto: {enc.get('nombre_del_producto','')}"])
    w.writerow([f"DCI: {enc.get('dci','')}"])
    w.writerow([f"Concentración: {enc.get('concentracion','')}"])
    w.writerow([f"Presentación: {enc.get('presentacion','')}"])
    w.writerow([f"Laboratorio: {enc.get('laboratorio','')}"])
    w.writerow([f"Rango: {d_desde} a {d_hasta}"])
    w.writerow([f"Almacén: #{alm}"])
    w.writerow([])

    headers = [
        "Fecha",
        "Cantidad Ingreso",
        "Cantidad Egreso",
        "Saldo Anterior",
        "Saldo Actual",
        "N° Receta",
        "Nombre del Paciente",
        "Nombre Médico",
        "N° de Factura",
        "Observaciones",
    ]
    w.writerow(headers)

    for r in rows:
        w.writerow(
            [
                r.get("fecha", ""),
                r.get("cantidad_ingreso", ""),
                r.get("cantidad_egreso", ""),
                r.get("saldo_anterior", ""),
                r.get("saldo_actual", ""),
                r.get("no_receta", ""),
                r.get("nombre_paciente", ""),
                r.get("nombre_medico", ""),
                r.get("nro_factura", ""),
                (r.get("observaciones", "") or "")
                .replace("\r", " ")
                .replace("\n", " "),
            ]
        )

    return resp


def export_kardex_xlsx(request):
    if not HAS_OPENPYXL:
        return HttpResponse(
            "Falta dependencia: openpyxl. Instala con: pip install openpyxl",
            status=501,
            content_type="text/plain; charset=utf-8",
        )

    try:
        enc, rows, d_desde, d_hasta, alm, med = _resolver_kardex_dataset(request)
    except Exception as e:
        return HttpResponse(
            f"Error: {e}", status=400, content_type="text/plain; charset=utf-8"
        )

    base = _build_med_label(enc, med)
    filename = f"KARDEX_{base}_DEL_{d_desde}_AL_{d_hasta}.xlsx"

    headers = [
        "Fecha",
        "Cantidad Ingreso",
        "Cantidad Egreso",
        "Saldo Anterior",
        "Saldo Actual",
        "N° Receta",
        "Nombre del Paciente",
        "Nombre Médico",
        "N° de Factura",
        "Observaciones",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kardex"

    # Título
    ws["A1"] = "KARDEX"
    ws["A1"].font = Font(bold=True, size=14)

    # Metadatos
    meta = [
        f"Producto: {enc.get('nombre_del_producto', '')}",
        f"DCI: {enc.get('dci', '')}",
        f"Concentración: {enc.get('concentracion', '')}",
        f"Presentación: {enc.get('presentacion', '')}",
        f"Laboratorio: {enc.get('laboratorio', '')}",
        f"Rango: {d_desde} a {d_hasta}",
        f"Almacén: #{alm}",
    ]
    row_idx = 2
    for line in meta:
        ws.cell(row=row_idx, column=1, value=line)
        row_idx += 1
    row_idx += 1  # blanco

    # Encabezados de tabla
    header_row = row_idx
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _coerce_excel_date(val):
        if (
            isinstance(val, str)
            and len(val) == 10
            and val[4] == "-"
            and val[7] == "-"
        ):
            try:
                return datetime.strptime(val, "%Y-%m-%d").date()
            except Exception:
                return val
        return val

    # Filas de datos
    for r in rows:
        row_idx += 1
        values = [
            _coerce_excel_date(r.get("fecha", "")),
            r.get("cantidad_ingreso", 0),
            r.get("cantidad_egreso", 0),
            r.get("saldo_anterior", 0),
            r.get("saldo_actual", 0),
            r.get("no_receta", ""),
            r.get("nombre_paciente", ""),
            r.get("nombre_medico", ""),
            r.get("nro_factura", ""),
            (r.get("observaciones", "") or "")
            .replace("\r", " ")
            .replace("\n", " "),
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Bordes
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(
        min_row=header_row,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(headers),
    ):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Formato de columnas
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(headers),
    ):
        for cell in row:
            col = cell.column
            if col == 1:
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = Alignment(horizontal="center")
            elif col in (2, 3, 4, 5):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col in (7, 8, 10):
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left")

    # Ancho de columnas (ligeramente ajustado)
    widths = [12, 14, 14, 14, 14, 14, 26, 26, 14, 30]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Congelar encabezado
    ws.freeze_panes = ws[f"A{header_row + 1}"]

    # ================== CONFIGURACIÓN DE IMPRESIÓN ==================
    # Orientación horizontal
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    # Ajustar a 1 página de ancho (alto libre)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    # Márgenes (en pulgadas)
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    # Repetir fila de encabezado de la tabla en cada página
    ws.print_title_rows = f"{header_row}:{header_row}"

    # Área de impresión: desde A1 hasta la última fila y última columna
    last_col_letter = get_column_letter(len(headers))
    last_row = ws.max_row
    ws.print_area = f"A1:{last_col_letter}{last_row}"
    # ================================================================

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def export_kardex_pdf(request):
    if not HAS_REPORTLAB:
        return HttpResponse(
            "Falta dependencia: reportlab. Instala con: pip install reportlab",
            status=501,
            content_type="text/plain; charset=utf-8",
        )

    try:
        enc, rows, d_desde, d_hasta, alm, med = _resolver_kardex_dataset(request)
    except Exception as e:
        return HttpResponse(
            f"Error: {e}", status=400, content_type="text/plain; charset=utf-8"
        )

    base = _build_med_label(enc, med)
    filename = f"KARDEX_{base}_DEL_{d_desde}_AL_{d_hasta}.pdf"

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        leftMargin=18,
        rightMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("<b>KARDEX</b>", styles["Title"]))
    meta_text = (
        f"Producto: {enc.get('nombre_del_producto','')} &nbsp;|&nbsp; "
        f"DCI: {enc.get('dci','')} &nbsp;|&nbsp; "
        f"Concentración: {enc.get('concentracion','')} &nbsp;|&nbsp; "
        f"Presentación: {enc.get('presentacion','')} &nbsp;|&nbsp; "
        f"Laboratorio: {enc.get('laboratorio','')} &nbsp;|&nbsp; "
        f"Rango: {d_desde} a {d_hasta} &nbsp;|&nbsp; "
        f"Almacén: #{alm}"
    )
    story.append(Paragraph(meta_text, styles["Normal"]))
    story.append(Spacer(1, 8))

    headers = [
        "Fecha",
        "Cant. Ingreso",
        "Cant. Egreso",
        "Saldo Ant.",
        "Saldo Act.",
        "N° Receta",
        "Nombre del Paciente",
        "Nombre Médico",
        "N° Factura",
        "Observaciones",
    ]
    data = [headers]
    for r in rows:
        data.append(
            [
                r.get("fecha", ""),
                r.get("cantidad_ingreso", ""),
                r.get("cantidad_egreso", ""),
                r.get("saldo_anterior", ""),
                r.get("saldo_actual", ""),
                r.get("no_receta", ""),
                r.get("nombre_paciente", ""),
                r.get("nombre_medico", ""),
                r.get("nro_factura", ""),
                (r.get("observaciones", "") or "")
                .replace("\r", " ")
                .replace("\n", " "),
            ]
        )

    col_widths = [
        2.3 * cm,
        2.8 * cm,
        2.8 * cm,
        2.8 * cm,
        2.8 * cm,
        3.2 * cm,
        6.8 * cm,
        6.8 * cm,
        3.0 * cm,
        7.0 * cm,
    ]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("ALIGN", (1, 1), (4, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f7f9f7")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )

    story.append(tbl)
    doc.build(story)

    bio.seek(0)
    resp = HttpResponse(bio.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
